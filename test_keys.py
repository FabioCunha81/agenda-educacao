import openpyxl, re, unicodedata

def clean(v):
    return re.sub(r'\s+', ' ', str(v or '')).strip()

def key(v):
    t = unicodedata.normalize('NFKD', clean(v)).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^A-Z0-9]+', ' ', t.upper()).strip()

wb = openpyxl.load_workbook('estatisticas.xlsx')
sheet = wb.active

for r in range(1, 22):
    print(key(sheet.cell(row=r, column=2).value) or key(sheet.cell(row=r, column=3).value))
